from typing import Tuple, List

import openpyxl
from lxml import etree as ET
from lxml.builder import E
from openpyxl.worksheet.worksheet import Worksheet


def create_ecert(row: Tuple) -> ET.Element:
    return E.ECERT(
        E.CERTNO(row[0]),
        E.CERTDATE(row[1].strftime("%Y-%m-%d")),
        E.STATUS(row[2]),
        E.IEC(f"0{row[3]}"),
        E.EXPNAME(f'"{row[4]}"'),
        E.BILLID(row[5]),
        E.SDATE(row[6].strftime("%Y-%m-%d")),
        E.SCC(row[7]),
        E.SVALUE(f"{row[8]}"),
    )


def get_data_rows(ws: Worksheet) -> List[Tuple]:
    return list(ws.iter_rows(min_row=6, values_only=True))


wb = openpyxl.load_workbook("test_input.xlsx", data_only=True)
ws = wb["Sheet1"]
filename = ws["B3"].value

base = ET.tostring(
    E.CERTDATA(
        E.FILENAME(filename),
        E.ENVELOPE(*[create_ecert(row) for row in get_data_rows(ws)]),
    ),
    pretty_print=True,
    encoding="unicode",
)

base = base.replace("  ", "\t")
formatted_xml = f'<?xml version="1.0" encoding="UTF-8"?>\n{base}'
with open("result_a.xml", "w") as f:
    f.write(formatted_xml)
