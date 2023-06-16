import requests
from bs4 import BeautifulSoup
from typing import Dict, Tuple, List
import openpyxl
from lxml.builder import E
from lxml import etree as ET
from datetime import datetime
from openpyxl.worksheet.worksheet import Worksheet

usd_rate_cache: Dict[str, float] = {}


def get_usd_rate(date: datetime) -> float:
    date_str = date.strftime("%d.%m.%Y")
    rate = usd_rate_cache.get(date_str)
    if rate is not None:
        return rate

    url = f"https://www.cbr.ru/currency_base/daily/?UniDbQuery.Posted=True&UniDbQuery.To={date_str}"
    response = requests.get(url)
    soup = BeautifulSoup(response.text, "html.parser")

    for tr in soup.find_all("tr"):
        tds = tr.find_all("td")
        if any("Доллар США" in td.get_text() for td in tds):
            rate = float(tds[4].get_text().replace(",", "."))
            usd_rate_cache[date_str] = rate
            return rate

    usd_rate_cache[date_str] = 1.0
    return usd_rate_cache[date_str]


def create_ecert(row: Tuple) -> ET.Element:
    return E.ECERT(
        E.CERTNO(row[0]),
        E.CERTDATE(row[1].strftime("%Y-%m-%d")),
        E.STATUS(row[2]),
        E.IEC(f"0{row[3]}"),
        E.EXPNAME(f'"{row[4]}"'),
        E.BILLID(row[5]),
        E.SDATE(row[6].strftime("%Y-%m-%d")),
        E.SCC(row[7]),
        E.SVALUE(f"{row[8]}"),
        E.SVALUEUSD(f"{row[8] / get_usd_rate(row[6]):.2f}"),
    )


def get_data_rows(ws: Worksheet) -> List[Tuple]:
    return list(ws.iter_rows(min_row=6, values_only=True))


wb = openpyxl.load_workbook("test_input.xlsx", data_only=True)
ws = wb["Sheet1"]
filename = ws["B3"].value

base = ET.tostring(
    E.CERTDATA(
        E.FILENAME(filename),
        E.ENVELOPE(*[create_ecert(row) for row in get_data_rows(ws)]),
    ),
    pretty_print=True,
    encoding="unicode",
)

base = base.replace("  ", "\t")
formatted_xml = f'<?xml version="1.0" encoding="UTF-8"?>\n{base}'
with open("result_b.xml", "w") as f:
    f.write(formatted_xml)
